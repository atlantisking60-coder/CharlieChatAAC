#!/usr/bin/env python3
"""AAC Board Compiler - dataclasses + multi-pass pipeline (incremental scaffold).

This file implements core dataclasses (`Tile`, `Board`, `AACProject`), a small
`AssetDatabase`, and a `AACBoardCompiler` that runs multiple passes. The parser
pass is conservative: it discovers headers and builds simple `VocabularyTile`
objects for each non-empty row. Link and asset resolution are implemented at a
basic level (case-insensitive asset lookup and board-name -> id mapping).
"""
from __future__ import annotations

import argparse
import dataclasses
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl
except Exception:
    openpyxl = None


def slugify(name: str) -> str:
    n = name.strip().lower()
    n = re.sub(r"[^a-z0-9]+", "_", n)
    n = re.sub(r"_+", "_", n)
    n = n.strip("_")
    return f"prebuilt_{n or 'board'}"


@dataclass
class Tile:
    id: str
    type: str
    label: Optional[str] = None
    image: Optional[str] = None
    linked_board_name: Optional[str] = None

    def to_dict(self) -> Dict:
        return {
            "id": self.id,
            "type": self.type,
            "label": self.label,
            "image": self.image,
            "linkedBoardName": self.linked_board_name,
        }


class VocabularyTile(Tile):
    def __init__(self, id: str, label: str, image: Optional[str] = None):
        super().__init__(id=id, type="vocabulary", label=label, image=image)


class BoardLinkTile(Tile):
    def __init__(self, id: str, label: str, linked_board_name: str):
        super().__init__(id=id, type="board_link", label=label, linked_board_name=linked_board_name)


@dataclass
class Board:
    id: str
    name: str
    columns: int = 6
    area: Optional[str] = None
    default_icon_folder: Optional[str] = None
    tiles: List[Tile] = field(default_factory=list)

    def to_dict(self) -> Dict:
        rows = max(1, (len(self.tiles) + self.columns - 1) // self.columns) if self.tiles else 1
        return {
            "id": self.id,
            "name": self.name,
            "area": self.area,
            "columns": self.columns,
            "defaultIconFolder": self.default_icon_folder,
            "layout": {
                "rows": rows,
                "blankTilesAdded": max(0, rows * self.columns - len(self.tiles)),
            },
            "tiles": [t.to_dict() for t in self.tiles],
        }


class AssetDatabase:
    def __init__(self, root: Path):
        self.root = root
        # index by filename and by relative path from root (both lowercase)
        self.index: Dict[str, Path] = {}
        self.index_rel: Dict[str, Path] = {}

    def scan(self):
        exts = {".png", ".jpg", ".jpeg", ".webp"}
        if not self.root.exists():
            return
        for p in self.root.rglob("*"):
            if p.suffix.lower() in exts:
                key = p.name.lower()
                rel = str(p.relative_to(self.root)).replace('\\', '/').lower()
                self.index[key] = p
                self.index_rel[rel] = p

    def find(self, name: str) -> Optional[Path]:
        if not name:
            return None
        key = Path(name).name.lower()
        # direct filename lookup
        if key in self.index:
            return self.index[key]
        # try with common extensions
        for ext in (".png", ".jpg", ".jpeg", ".webp"):
            k = key + ext
            if k in self.index:
                return self.index[k]
        return None

    def find_in_folder(self, folder: str, name: str) -> Optional[Path]:
        if not folder:
            return self.find(name)
        folder_norm = folder.strip().lstrip('./').rstrip('/')
        candidate = f"{folder_norm}/{Path(name).name}".lower()
        if candidate in self.index_rel:
            return self.index_rel[candidate]
        # try with extensions
        for ext in (".png", ".jpg", ".jpeg", ".webp"):
            c2 = candidate + ext
            if c2 in self.index_rel:
                return self.index_rel[c2]
        # fallback to name lookup
        return self.find(name)


@dataclass
class AACProject:
    boards: Dict[str, Board] = field(default_factory=dict)
    assets: AssetDatabase | None = None
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def validate_project(project: AACProject) -> (List[str], List[str]):
    """Perform validation checks on the project and return (errors, warnings).

    Checks:
    - duplicate board ids
    - duplicate tile ids
    - empty boards
    - circular board link detection
    - JSON serializability of each board
    """
    errors: List[str] = []
    warnings: List[str] = []

    # duplicate board IDs
    id_counts: Dict[str, int] = {}
    for b in project.boards.values():
        id_counts[b.id] = id_counts.get(b.id, 0) + 1
    for bid, cnt in id_counts.items():
        if cnt > 1:
            errors.append(f"Duplicate board id: {bid}")

    # duplicate tile ids
    tile_ids: Dict[str, int] = {}
    for b in project.boards.values():
        for t in b.tiles:
            tile_ids[t.id] = tile_ids.get(t.id, 0) + 1
    for tid, cnt in tile_ids.items():
        if cnt > 1:
            errors.append(f"Duplicate tile id: {tid}")

    # empty boards
    for b in project.boards.values():
        if len(b.tiles) == 0:
            warnings.append(f"Empty board: {b.name}")

    # build graph of board links (assume linked_board_name contains board id after resolution)
    graph: Dict[str, List[str]] = {}
    for b in project.boards.values():
        graph[b.id] = []
        for t in b.tiles:
            if isinstance(t, BoardLinkTile) and t.linked_board_name:
                graph[b.id].append(str(t.linked_board_name))

    # detect cycles using DFS
    visited: Dict[str, int] = {}  # 0=unvisited,1=visiting,2=done

    def dfs(node: str) -> bool:
        state = visited.get(node, 0)
        if state == 1:
            return True
        if state == 2:
            return False
        visited[node] = 1
        for neigh in graph.get(node, []):
            if neigh not in graph:
                continue  # already reported as warning in pass 4
            if dfs(neigh):
                return True
        visited[node] = 2
        return False

    for node in graph:
        if dfs(node):
            warnings.append("Circular board link detected involving: " + node)
            break

    # json serializable check
    for b in project.boards.values():
        try:
            json.dumps(b.to_dict())
        except Exception as e:
            errors.append(f"Board not JSON serializable: {b.name} -> {e}")

    return errors, warnings


class AACBoardCompiler:
    def __init__(self, workbook: Path, out_dir: Path, asset_root: Optional[Path] = None, verbose: bool = False):
        self.workbook = workbook
        self.out_dir = out_dir
        self.verbose = verbose
        self.asset_root = asset_root or Path("assets")
        self.project = AACProject()
        self.log_path = Path("logs") / "compiler.log"
        self.log_path.parent.mkdir(parents=True, exist_ok=True)

    def log(self, *args):
        msg = " ".join(str(a) for a in args)
        if self.verbose:
            print(msg)
        with self.log_path.open("a", encoding="utf-8") as fh:
            fh.write(msg + "\n")

    def run(self, check: bool = False) -> int:
        if openpyxl is None:
            msg = "Error: openpyxl is required. Install from requirements.txt or pip install openpyxl"
            print(msg)
            with self.log_path.open("a", encoding="utf-8") as fh:
                fh.write(msg + "\n")
            return 2

        self.log("Loading workbook:", self.workbook)
        wb = openpyxl.load_workbook(self.workbook, read_only=True, data_only=True)

        # Pass 1: scan all sheets, discover headers, create one Board per data row
        # (BOARD NAME column drives board identity, not sheet tab name)

        def split_list(cell_val) -> List[str]:
            if not cell_val:
                return []
            s = str(cell_val)
            parts = re.split(r"[,;\n]+", s)
            return [p.strip() for p in parts if p.strip()]

        class ImageViewerTile(Tile):
            def __init__(self, id: str, label: Optional[str], image: str):
                super().__init__(id=id, type="image_viewer", label=label, image=image)

        def is_placeholder(name: str) -> bool:
            """Return True for instruction/note text that is not a real board name."""
            s = name.strip()
            return s.startswith("(") or s.startswith("[") or len(s) == 0

        # First sub-pass: collect all board names so links can be resolved
        raw_board_rows: List[Dict] = []  # list of {sheet, headers, header_lowers, idx_*, row}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers: List[str] = []
            header_row_num = None
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    headers = [str(cell).strip() if cell is not None else "" for cell in row]
                    header_row_num = r_idx
                    break
            if not headers:
                continue

            header_lowers = [h.lower() for h in headers]

            def idx_for(substrings: List[str], hl=header_lowers) -> Optional[int]:
                for i, h in enumerate(hl):
                    for sub in substrings:
                        if sub in h:
                            return i
                return None

            idx_area = idx_for(["area"])
            idx_board_name = idx_for(["board name"])
            idx_columns = idx_for(["columns"])
            idx_words = idx_for(["words on board", "words"])
            idx_sub_start = idx_for(["subboards at start"])
            idx_sub_end = idx_for(["subboards at end"])
            idx_click_view = idx_for(["click to open full size picture", "click to open"])
            idx_folder_word_icons = idx_for(["folder path for word icons", "word icon"])
            idx_folder_link_icons = idx_for(["folder path for folder link icons", "folder link"])

            start_row = (header_row_num or 1) + 1
            for r in ws.iter_rows(min_row=start_row, values_only=True):
                if not any(cell is not None and str(cell).strip() != "" for cell in r):
                    continue
                raw_board_rows.append({
                    "sheet": sheet_name,
                    "r": r,
                    "idx_area": idx_area,
                    "idx_board_name": idx_board_name,
                    "idx_columns": idx_columns,
                    "idx_words": idx_words,
                    "idx_sub_start": idx_sub_start,
                    "idx_sub_end": idx_sub_end,
                    "idx_click_view": idx_click_view,
                    "idx_folder_word_icons": idx_folder_word_icons,
                    "idx_folder_link_icons": idx_folder_link_icons,
                })

        # Pass 1b: create Board objects keyed by display name
        for entry in raw_board_rows:
            r = entry["r"]
            idx_board_name = entry["idx_board_name"]
            if idx_board_name is None:
                continue
            raw_name = r[idx_board_name]
            if raw_name is None:
                continue
            name = str(raw_name).strip()
            if not name or is_placeholder(name):
                continue
            if name not in self.project.boards:
                board_id = slugify(name)
                board = Board(id=board_id, name=name)
                self.project.boards[name] = board
                self.log("Discovered board:", name, "->", board_id)

        # Pass 2: build asset database
        self.project.assets = AssetDatabase(self.asset_root)
        self.project.assets.scan()
        self.log(f"Assets indexed: {len(self.project.assets.index)}")

        # Pass 3: parse rows into tiles
        for entry in raw_board_rows:
            r = entry["r"]
            idx_board_name = entry["idx_board_name"]
            idx_area = entry["idx_area"]
            idx_columns = entry["idx_columns"]
            idx_words = entry["idx_words"]
            idx_sub_start = entry["idx_sub_start"]
            idx_sub_end = entry["idx_sub_end"]
            idx_click_view = entry["idx_click_view"]
            idx_folder_word_icons = entry["idx_folder_word_icons"]

            if idx_board_name is None:
                continue
            raw_name = r[idx_board_name]
            if raw_name is None:
                continue
            name = str(raw_name).strip()
            if not name or is_placeholder(name):
                continue
            board = self.project.boards.get(name)
            if board is None:
                continue

            board_counter = len(board.tiles)

            if idx_area is not None and r[idx_area]:
                board.area = str(r[idx_area]).strip()
            if idx_columns is not None:
                try:
                    val = r[idx_columns]
                    if val is not None and str(val).strip().isdigit():
                        board.columns = int(str(val).strip())
                except Exception:
                    pass
            if idx_folder_word_icons is not None and r[idx_folder_word_icons]:
                board.default_icon_folder = str(r[idx_folder_word_icons]).strip()

            if idx_sub_start is not None and r[idx_sub_start]:
                for link_name in split_list(r[idx_sub_start]):
                    if is_placeholder(link_name):
                        continue
                    board_counter += 1
                    tid = f"{board.id}_t{board_counter}"
                    board.tiles.append(BoardLinkTile(id=tid, label=link_name, linked_board_name=link_name))

            if idx_words is not None and r[idx_words]:
                folder = str(r[idx_folder_word_icons]).strip() if idx_folder_word_icons is not None and r[idx_folder_word_icons] else None
                for w in split_list(r[idx_words]):
                    board_counter += 1
                    tid = f"{board.id}_t{board_counter}"
                    image_val = f"{folder}/{w}" if folder else None
                    board.tiles.append(VocabularyTile(id=tid, label=w, image=image_val))

            if idx_click_view is not None and r[idx_click_view]:
                for img in split_list(r[idx_click_view]):
                    board_counter += 1
                    tid = f"{board.id}_t{board_counter}"
                    board.tiles.append(ImageViewerTile(id=tid, label=None, image=img))

            if idx_sub_end is not None and r[idx_sub_end]:
                for link_name in split_list(r[idx_sub_end]):
                    if is_placeholder(link_name):
                        continue
                    board_counter += 1
                    tid = f"{board.id}_t{board_counter}"
                    board.tiles.append(BoardLinkTile(id=tid, label=link_name, linked_board_name=link_name))

        # Pass 4: resolve board links (map names to ids)
        name_to_id = {b.name.lower(): b.id for b in self.project.boards.values()}
        name_to_id.update({slugify(b.name).lower(): b.id for b in self.project.boards.values()})
        for b in self.project.boards.values():
            for t in b.tiles:
                if isinstance(t, BoardLinkTile) and t.linked_board_name:
                    raw_name = str(t.linked_board_name).strip()
                    target = name_to_id.get(raw_name.lower())
                    if not target:
                        target = next((board.id for board in self.project.boards.values() if board.name.lower() == raw_name.lower()), None)
                    if not target:
                        self.project.warnings.append(f"Unresolved board link '{t.linked_board_name}' in board '{b.name}'")
                    else:
                        t.linked_board_name = target

        # Pass 5: resolve assets for tiles
        images_found = 0
        images_missing = 0
        for b in self.project.boards.values():
            for t in b.tiles:
                if t.image:
                    # t.image may be like 'folder/name' or just 'name'
                    img = None
                    if self.project.assets:
                        # try folder-aware lookup
                        if isinstance(t.image, str) and '/' in t.image:
                            parts = t.image.replace('\\', '/').split('/')
                            folder = '/'.join(parts[:-1])
                            name = parts[-1]
                            img = self.project.assets.find_in_folder(folder, name)
                        else:
                            img = self.project.assets.find(str(t.image))
                    if img:
                        images_found += 1
                        # store path relative to asset root for readability
                        try:
                            t.image = str(img.relative_to(self.project.assets.root))
                        except Exception:
                            t.image = str(img)
                    else:
                        images_missing += 1
                        self.project.warnings.append(f"Missing image '{t.image}' for tile '{t.id}' on board '{b.name}'")

        # Pass 6: validation - duplicates, empty boards, invalid IDs
        # duplicate board IDs
        id_counts: Dict[str, int] = {}
        for b in self.project.boards.values():
            id_counts[b.id] = id_counts.get(b.id, 0) + 1
        for bid, cnt in id_counts.items():
            if cnt > 1:
                self.project.errors.append(f"Duplicate board id: {bid}")

        # duplicate tile ids
        tile_ids: Dict[str, int] = {}
        for b in self.project.boards.values():
            for t in b.tiles:
                tile_ids[t.id] = tile_ids.get(t.id, 0) + 1
        for tid, cnt in tile_ids.items():
            if cnt > 1:
                self.project.errors.append(f"Duplicate tile id: {tid}")

        # empty boards
        for b in self.project.boards.values():
            if len(b.tiles) == 0:
                self.project.warnings.append(f"Empty board: {b.name}")

        # collect statistics
        stats = {
            "boards": len(self.project.boards),
            "tiles": sum(len(b.tiles) for b in self.project.boards.values()),
            "images_found": images_found,
            "images_missing": images_missing,
            "errors": len(self.project.errors),
            "warnings": len(self.project.warnings),
        }

        self.log("Stats:", stats)

        # shift pass numbers for remaining operations
        # Pass 7: layout engine (ensure rectangular layout by adding blanks)

        
        # (layout engine follows)
        for b in self.project.boards.values():
            total = len(b.tiles)
            cols = b.columns or 6
            rows = (total + cols - 1) // cols
            needed = rows * cols - total
            for i in range(needed):
                bid = f"{b.id}_blank_{i}"
                b.tiles.append(Tile(id=bid, type="blank"))

        # Pass 6: validation using shared helper
        errors, warnings = validate_project(self.project)
        self.project.errors.extend(e for e in errors if e not in self.project.errors)
        self.project.warnings.extend(w for w in warnings if w not in self.project.warnings)
        # Only hard errors (duplicate IDs, JSON serialization failures) abort the build
        hard_errors = [e for e in self.project.errors if not e.startswith("Unresolved")]
        if hard_errors:
            for e in hard_errors:
                print("ERROR:", e)
            print(f"Aborting: {len(hard_errors)} error(s) found.")
            return 3

        # Pass 7: write JSON
        self.out_dir.mkdir(parents=True, exist_ok=True)
        for b in self.project.boards.values():
            if b.area:
                area_folder = self.out_dir / b.area.strip().lower()
            else:
                area_folder = self.out_dir / "uncategorised"
            area_folder.mkdir(parents=True, exist_ok=True)
            out_file = area_folder / f"{b.id}.json"
            with open(out_file, "w", encoding="utf-8") as f:
                json.dump(b.to_dict(), f, indent=2, ensure_ascii=False)
            self.log("Wrote board:", out_file)

        self.log("Warnings:", len(self.project.warnings))
        self.log("Compiler completed. See", self.log_path)
        return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", "-w", default="Board Structure.xlsx")
    parser.add_argument("--out", "-o", default="lib/data/boards")
    parser.add_argument("--asset-root", default="assets")
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    compiler = AACBoardCompiler(workbook=Path(args.workbook), out_dir=Path(args.out), asset_root=Path(args.asset_root), verbose=args.verbose)
    return compiler.run(check=args.check)


if __name__ == "__main__":
    raise SystemExit(main())
